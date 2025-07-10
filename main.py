import openpyxl

# Load workbook and sheet
file_path = 'Template.xlsx' 
wb = openpyxl.load_workbook(file_path)
ws = wb.active
ws.title = "RMC4916_Example"  # Set sheet name

# Constants
start_row = 13
aj_column = 36  # AJ
ai_column = 35  # AI
description_columns = []

# Step 1: Identify "Description / Content" columns and map to product names

col = 3  
while col <= 34: 
    header_12 = ws.cell(row=12, column=col + 1).value  
    if header_12 and "Description / Content" in str(header_12):
        product_name = ws.cell(row=11, column=col).value 
        description_columns.append((col + 1, product_name))
    col += 2  # Move to next product block (2 columns per product)

# Step 2: Loop through each data row

for row in range(start_row, ws.max_row + 1):
    matching_products = []
    found_no_trace = False

    for col_index in range(3, 35): 
        cell_value = ws.cell(row=row, column=col_index).value
        if isinstance(cell_value, str) and "No Trace" in cell_value:
            found_no_trace = True

    for col_index, product in description_columns:
        cell_value = ws.cell(row=row, column=col_index).value
        if isinstance(cell_value, str) and "No Trace" in cell_value:
            matching_products.append(product)

    if found_no_trace:
        ws.cell(row=row, column=ai_column).value = "N"
    if matching_products:
        ws.cell(row=row, column=aj_column).value = ", ".join(matching_products)

# Save the workbook

updated_file = 'updated_' + file_path
wb.save(updated_file)
print(f"Update complete. Saved as '{updated_file}'")
